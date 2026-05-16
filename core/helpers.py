import exceptions
from core.logger import get_logger

logger = get_logger()

import os
import datetime
import pandas as pd
import random
import openpyxl
from openpyxl.styles import Font
from time import perf_counter

def salvar_em_planilha_excel(df, nome_coleta: str, sufixo_data: str = None):
    """
    Exporta um DataFrame do Pandas para um arquivo Excel (.xlsx) na pasta 'Downloads' do usuário.
    Gera um nome de arquivo único com base no nome_coleta, sufixo_data e um código numérico aleatório.
    """
    if df is not None and not df.empty:
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        
        if not sufixo_data:
            sufixo_data = datetime.datetime.today().strftime('%d_%m_%Y')
        else:
            sufixo_data = sufixo_data.replace("-", "_").replace("/", "_")
            
        nome_arquivo = f"{nome_coleta}_{sufixo_data}_cod{random.randint(1000, 9999)}.xlsx"
        caminho_final = os.path.join(downloads_path, nome_arquivo)

        try:
            df.to_excel(caminho_final, index=False)
        except Exception as e:
            logger.error(f"Dev Error: Falha ao salvar Excel: {e}", exc_info=True)
            raise exceptions.DataCleaningException("Erro ao salvar a planilha Excel. Verifique se o arquivo não está aberto em outro programa.")
    else:
        print("Nenhum dado encontrado.")
    return

def atualizar_planilha_produtividade(df, caminho_planilha):
    """
    Atualiza a planilha original inserindo os dados de forma incremental na aba 'Base em min'
    e adicionando as fórmulas nas colunas não mapeadas diretamente.
    """
    if df is None or df.empty:
        print("Nenhum dado para atualizar.")
        return

    try:
        meses_pt = {1:'janeiro', 2:'fevereiro', 3:'março', 4:'abril', 5:'maio', 6:'junho', 
                    7:'julho', 8:'agosto', 9:'setembro', 10:'outubro', 11:'novembro', 12:'dezembro'}
        
        # Garante que 'Data' seja datetime
        df['Data'] = pd.to_datetime(df['Data'])
        
        # Calcula Mês Anterior e Ano de uma vez só
        df['mes_idx'] = df['Data'].dt.month - 1
        df.loc[df['mes_idx'] == 0, 'mes_idx'] = 12
        df['mes_anterior'] = df['mes_idx'].map(meses_pt)
        df['ano_val'] = df['Data'].dt.year

        # 2. Carga do Workbook
        start_loading = perf_counter()
        wb = openpyxl.load_workbook(caminho_planilha)
        print(f"O lxml está sendo usado? {openpyxl.LXML}")
        end_loading = perf_counter()
        print(f"Tempo de carga: {end_loading - start_loading:.4f} segundos")
        
        start_append = perf_counter()
        try:
            ws = wb["Base em min"]
            start_row = ws.max_row + 1
            font_8 = Font(size=8)

            # 3. Iteração Otimizada com ws.append (levando em média 90 segundos, da pra otimizar)
            colunas_uso = ['Data', 'Operador', 'Caso', 'Cliente', 'Estação', 'Esteira', 'Tempo_trabalhado_min', 'mes_anterior', 'ano_val']
            
            for row in df[colunas_uso].itertuples(index=False):
                # Calcular os dados e fórmulas antecipadamente como uma lista
                nova_linha = [
                    row.Data, row.Operador, row.Caso, row.Cliente, row.Estação, row.Esteira, row.Tempo_trabalhado_min, row.mes_anterior,
                    f"=VLOOKUP(F{start_row},Infor!$A:$B,2,0)",
                    f'=I{start_row}&"-"&E{start_row}',
                    f"=VLOOKUP(J{start_row},'Faixas de trabalho'!$D:$E,2,0)",
                    f"=G{start_row}/60",
                    f"=VLOOKUP(J{start_row},'Faixas de trabalho'!$D:$G,4,0)",
                    row.ano_val,
                    f"=VLOOKUP(J{start_row},'Faixas de trabalho'!$D:$F,3,0)"
                ]
                
                # Append insere a linha de uma vez só com todos os dados
                ws.append(nova_linha)
                
                # Otimização: Acessa a tupla de células da linha criada de uma vez só
                celulas_linha = ws[start_row]
                
                celulas_linha[0].number_format = 'dd/mm/yyyy' # Coluna 1
                celulas_linha[6].number_format = '0'          # Coluna 7
                celulas_linha[11].number_format = '0.00'      # Coluna 12
                
                # Aplica a fonte sem precisar fazer lookup de célula individualmente
                for i in range(15):
                    celulas_linha[i].font = font_8
                    
                start_row += 1
            end_append = perf_counter()
            print(f"Tempo de append: {end_append - start_append:.4f} segundos")
            
            # 4. Salvar apenas UMA VEZ ao final
            start_save = perf_counter()
            wb.save(caminho_planilha)
            end_save = perf_counter()
            print(f"Tempo de salvamento: {end_save - start_save:.4f} segundos")
        finally:
            if 'wb' in locals():
                wb.close()
        
    except Exception as e:
        logger.error(f"Dev Error: Falha ao atualizar planilha: {e}", exc_info=True)
        raise exceptions.DataCleaningException(f"Erro ao atualizar a planilha. Verifique se o arquivo não está aberto e se a aba 'Base em min' existe. Detalhe: {e}")

def limpeza_dados(coleta):
    """Converte a lista de dicionários brutos em um DataFrame estruturado do Pandas."""
    df = pd.DataFrame(coleta)
    return df

def limpeza_operadores_decorator(f):
    """
    Decorator que intercepta o DataFrame de operadores gerado e aplica formatações rígidas.
    Ajusta datas para o formato local e converte colunas numéricas antes do salvamento.
    """
    def wrapper(coleta):
        df = f(coleta)
        try:
            try:
                df["Data"] = df["Data"].str.replace("-", "/")
            except ValueError as e:
                logger.warning(f"Aviso na limpeza de Data (Operadores): {e}")

            df["Data"] = pd.to_datetime(df["Data"], dayfirst=True)
            df["Tempo_trabalhado_min"] = pd.to_numeric(df["Tempo_trabalhado_min"], errors="coerce")
            df["Caso"] = pd.to_numeric(df["Caso"], errors="coerce")
            return df
        except Exception as e:
            logger.error(f"Dev Error: Falha na limpeza de Operadores: {e}", exc_info=True)
            raise exceptions.DataCleaningException("Falha ao organizar colunas da tabela de operadores. O formato no site do Mapia pode ter mudado.")

    return wrapper

def limpeza_marcos_decorator(f):
    """
    Decorator que intercepta o DataFrame de marcos gerado e aplica formatações rígidas.
    Trunca datas para extrair apenas o dia e converte strings contábeis em tipos numéricos.
    """
    def wrapper(coleta):
        df_marcos = f(coleta)
        try:
            df_marcos["Data Alcan"] = df_marcos["Data Alcan"].str.slice(0, 10).replace("-", "/")
            df_marcos["Data Alcan"] = pd.to_datetime(df_marcos["Data Alcan"], dayfirst=True)

            df_marcos["Caso"] = pd.to_numeric(df_marcos["Caso"], errors="coerce")
            df_marcos["Obrigações Totais"] = pd.to_numeric(df_marcos["Obrigações Totais"], errors="coerce")
            df_marcos["Obrigações Produzidas"] = pd.to_numeric(df_marcos["Obrigações Produzidas"], errors="coerce")
            return df_marcos
        except Exception as e:
            logger.error(f"Dev Error: Falha na limpeza de Marcos: {e}", exc_info=True)
            raise exceptions.DataCleaningException("Falha ao organizar colunas da tabela de marcos. O formato no site do Mapia pode ter mudado.")

    return wrapper

@limpeza_operadores_decorator
def limpar_dados_operadores(coleta):
    """Aplica a limpeza geral e sobrescreve com as regras de operadores."""
    return limpeza_dados(coleta)

@limpeza_marcos_decorator
def limpar_dados_marcos(coleta):
    """Aplica a limpeza geral e sobrescreve com as regras de marcos."""
    return limpeza_dados(coleta)
